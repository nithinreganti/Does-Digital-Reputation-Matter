import openpyxl
from selenium import webdriver
import time
from tempfile import TemporaryFile
import csv
from statistics import mean
import numpy as np
start_time=time.time()
path = "F:\\erin\\urls_twitter.xlsx"
wb_obj = openpyxl.load_workbook(path) 
sheet_obj = wb_obj.active 
m_row = sheet_obj.max_row
driver=webdriver.Chrome() 

Total_tweets=[]
total_followers=[]
Total_following=[]
Total_likes=[]
max_likes=[]
min_likes=[]



for i in range(37,38):#m_row + 1): 
            cell_obj = sheet_obj.cell(row = i, column = 2) 
            try:
                      
            
                twitter_url=cell_obj.value 
                if 'null' in twitter_url:
                   Total_tweets.append(0) 
                   total_followers.append(0)
                   Total_following.append(0)
                   Total_likes.append(0)
                   max_likes.append(0)
                   min_likes.append(0)
                else:                   
                    driver.get(twitter_url)
                    time.sleep(5)
                    try:
                        tweets=driver.find_element_by_xpath('//*[@id="page-container"]/div[1]/div/div[2]/div/div/div[2]/div/div/ul/li[1]/a/span[3]')
                        tweet_count=tweets.text
                        if 'K' in tweet_count:
                            tweet_count_string=tweet_count.replace('K','')
                            tweet_count_integer=float(tweet_count_string)
                            tweet_count=tweet_count_integer* 1000                    
                        Total_tweets.append(tweet_count)
                        followers=driver.find_element_by_xpath('//*[@id="page-container"]/div[1]/div/div[2]/div/div/div[2]/div/div/ul/li[3]/a/span[3]')
                        follower_count=followers.text
                        if 'K' in follower_count:
                            follower_count_string=follower_count.replace('K','')
                            follower_count_integer=float(follower_count_string)
                            follower_count=follower_count_integer* 1000
                        total_followers.append(follower_count)
                        following=driver.find_element_by_xpath('//*[@id="page-container"]/div[1]/div/div[2]/div/div/div[2]/div/div/ul/li[2]/a/span[3]')
                        following_count=following.text
                        if 'K' in following_count:
                            following_count_string=following_count.replace('K','')
                            following_count_integer=float(following_count_string)
                            following_count=following_count_integer* 1000 
                        Total_following.append(following_count)
                    except:
                          Total_tweets.append(0) 
                          total_followers.append(0)
                          Total_following.append(0)
                    try:       
                        likes=driver.find_element_by_xpath('//*[@id="page-container"]/div[1]/div/div[2]/div/div/div[2]/div/div/ul/li[4]/a/span[3]')
                        likes_count=likes.text
                        if 'K' in likes_count:
                            likes_count_string=likes_count.replace('K','')
                            likes_count_integer=float(likes_count_string)
                            likes_count=likes_count_integer* 1000                    
                        Total_likes.append(likes_count)
                        SCROLL_PAUSE_TIME = 0.5
                        last_height = driver.execute_script("return document.body.scrollHeight")                    
                        while True:
                            # Scroll down to bottom
                            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                        
                            # Wait to load page
                            time.sleep(SCROLL_PAUSE_TIME)
                        
                            # Calculate new scroll height and compare with last scroll height
                            new_height = driver.execute_script("return document.body.scrollHeight")
                            if new_height == last_height:
                                break
                            last_height = new_height
                        time.sleep(10)        
                        Latest=driver.find_elements_by_class_name('ProfileTweet-actionCountForPresentation')
                        latest_list=[]
                        for i in Latest:
                            latest_list.append(i.text)
                            
                        max_likes.append(max(latest_list))
                        min_temp=min(latest_list)                  
                        if '' in  min_temp:
                            min_likes.append(0)
                    except:
                        max_likes.append(0)
                        min_likes.append(0)
                        Total_likes.append(0)
                    
      
            except:
                 print('e')
                 
                 Total_tweets.append('null') 
                 total_followers.append('null') 
                 Total_following.append('null') 
                 Total_likes.append('null')
                 max_likes.append('null')
                 min_likes.append('null')
                 
                 
                 
rows = zip(Total_tweets,total_followers,Total_following,Total_likes,max_likes,min_likes)
with open('ranking_twitter.csv', "w",newline='') as f:
    writer = csv.writer(f)
    rowheader = ["Tweet_count","Followers","Following","Total_likes","Max_likes","Minimum_likes"]
    writer.writerow(rowheader) 
    for row in rows:
        writer.writerow(row)

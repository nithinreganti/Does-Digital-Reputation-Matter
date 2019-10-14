import openpyxl
from selenium import webdriver
import time
from statistics import mean
import xlwt
from tempfile import TemporaryFile
import csv

start_time=time.time()
path = "F:\\erin\\urls_youtube.xlsx"
wb_obj = openpyxl.load_workbook(path) 
sheet_obj = wb_obj.active 
m_row = sheet_obj.max_row

chrome_options = webdriver.ChromeOptions()
time.sleep(5)
prefs = {"profile.default_content_setting_values.notifications" : 2}
chrome_options.add_experimental_option("prefs",prefs)
driver = webdriver.Chrome(chrome_options=chrome_options)

Total_subscribers=[]
Total_views=[]
Maximum_views=[]
Average_views=[]
Minimum_views=[]

for i in range(184,185):# m_row + 1): 
            cell_obj = sheet_obj.cell(row = i, column = 2) 
            try:
                
                Youtube_url=cell_obj.value
                if 'null' in Youtube_url:
                       Total_subscribers.append(0) 
                       Total_views.append(0) 
                       Maximum_views.append(0) 
                       Average_views.append(0)
                       Minimum_views.append(0)
                       
                else:
                
                        url=Youtube_url
                        driver.get(url)
                        time.sleep(10)
                        try:
                            subscribers=driver.find_element_by_xpath('//*[@id="subscriber-count"]')                                                                 
                            subscribers_text=subscribers.text
                            subscribers_count=subscribers_text.replace('subscribers','')                       
                         
                            if 'K' in subscribers_count:
                                    subscribers_count_string=subscribers_count.replace('K','')
                                    subscribers_count_integer=float(subscribers_count_string)
                                    subscribers_count=subscribers_count_integer* 1000 
                            Total_subscribers.append(subscribers_count)
                        except:
                            Total_subscribers.append(0)
                            print('couldnt fetch subsribers')
                            
                        try:
                            x=driver.find_element_by_xpath('//*[@id="tabsContent"]/paper-tab[2]')
                            x.click()
                            time.sleep(5)                        
                            SCROLL_PAUSE_TIME = 0.5
                            last_height = driver.execute_script("return document.body.scrollHeight")                    
                            while True:
                                # Scroll down to bottom
                                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                            
                                # Wait to load page
                                time.sleep(SCROLL_PAUSE_TIME)
                            
                                # Calculate new scroll height and compare with last scroll height
                                new_height = driver.execute_script("return document.body.scrollHeight")
                                if new_height == last_height:
                                    break
                                last_height = new_height
                            time.sleep(10)          
                            
                            views=driver.find_elements_by_css_selector('#metadata-line > span:nth-child(1)')
                                                                       
                            view_count_individual=[]                                  
                            for i in (views):   
                                    if 'K' in i.text :
                                        i_string=i.text.replace('K views','')       
                                        i_integer=float(i_string)
                                        i_count=i_integer* 1000
                                        view_count_individual.append(i_count)
                                    elif 'views' in i.text:
                                        i_string=i.text.replace('views','')
                                        if 'No' in i_string:
                                            view_count_individual.append(0)
                                        else:        
                                            i_integer=float(i_string)
                                            view_count_individual.append(i_integer)
                           
                            Total_views.append(sum(view_count_individual))
                            Maximum_views.append(max(view_count_individual))
                            Average_views.append(mean(view_count_individual))
                            Minimum_views.append(min(view_count_individual))
                        except:
                            
                            Total_views.append(0)
                            Maximum_views.append(0)
                            Average_views.append(0)
                            Minimum_views.append(0)
                            
                                                            
            
            except:
                print('error')
                
                Total_subscribers.append('null') 
                Total_views.append('null') 
                Maximum_views.append('null') 
                Average_views.append('null')
                Minimum_views.append('null')
                
rows = zip(Total_subscribers,Total_views,Maximum_views,Average_views)
with open('rankingyoutube.csv', "w",newline='') as f:
    writer = csv.writer(f)
    rowheader = ["Subscribers","Total views","Maximum views","Average views"]
    writer.writerow(rowheader) 
    for row in rows:
        writer.writerow(row)
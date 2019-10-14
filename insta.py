import openpyxl
from selenium import webdriver
import time
from statistics import mean
import xlwt
from tempfile import TemporaryFile
import csv

start_time=time.time()
path = "F:\\erin\\urls_insta.xlsx"
wb_obj = openpyxl.load_workbook(path) 
sheet_obj = wb_obj.active 
m_row = sheet_obj.max_row

chrome_options = webdriver.ChromeOptions()
time.sleep(5)
prefs = {"profile.default_content_setting_values.notifications" : 2}
chrome_options.add_experimental_option("prefs",prefs)
driver = webdriver.Chrome(chrome_options=chrome_options)
driver.get('https://www.instagram.com')
time.sleep(5)
switchaccount=driver.find_element_by_xpath('//*[@id="react-root"]/section/main/article/div[2]/div[2]/p/a')
switchaccount.click()
time.sleep(5)
username=driver.find_element_by_xpath('//*[@id="react-root"]/section/main/div/article/div/div[1]/div/form/div[2]/div/label/input')
username.send_keys('abcdx6667@gmail.com')
password=driver.find_element_by_xpath('//*[@id="react-root"]/section/main/div/article/div/div[1]/div/form/div[3]/div/label/input')
password.send_keys('qwerty1234')
login_button=driver.find_element_by_xpath('//*[@id="react-root"]/section/main/div/article/div/div[1]/div/form/div[4]/button/div')
login_button.click()


Total_posts=[]
Total_followers=[]
Total_following=[]
max_likes=[]
    
for i in range(150,218):# m_row + 1): 
            cell_obj = sheet_obj.cell(row = i, column = 2) 
            try:
                insta_url=cell_obj.value
                if 'null' in insta_url:
                       Total_posts.append(0) 
                       Total_followers.append(0) 
                       Total_following.append(0) 
                else:                
                 
                    driver.get(insta_url)
                      
                    time.sleep(10) 
                    SCROLL_PAUSE_TIME = 15 
                    try: 
                        posts=driver.find_element_by_xpath('//*[@id="react-root"]/section/main/div/header/section/ul/li[1]/a/span')
                        Post_count=posts.text               
                        Total_posts.append(Post_count)
                        Followers=driver.find_element_by_xpath('//*[@id="react-root"]/section/main/div/header/section/ul/li[2]/a/span')
                        Follower_count=Followers.text
                        if 'k' in Follower_count:
                                follower_count_string=Follower_count.replace('k','')
                                follower_count_integer=float(follower_count_string)
                                Follower_count=follower_count_integer* 1000
                        Total_followers.append(Follower_count)
                        Following=driver.find_element_by_xpath('//*[@id="react-root"]/section/main/div/header/section/ul/li[3]/a/span')
                        Following_count=Following.text
                        Total_following.append(Following_count)
                    except:
                        Total_posts.append(0)
                        Total_followers.append(0)
                        Total_following.append(0)
                    '''
                    SCROLL_PAUSE_TIME = 0.5
                    last_height = driver.execute_script("return document.body.scrollHeight")                    
                    while True:
                        # Scroll down to bottom
                        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    
                        # Wait to load page
                        time.sleep(SCROLL_PAUSE_TIME)
                    
                        # Calculate new scroll height and compare with last scroll height
                        new_height = driver.execute_script("return document.body.scrollHeight")
                        if new_height == last_height:
                            break
                        last_height = new_height
                    time.sleep(5)
                    
                    likes=driver.find_elements_by_class_name('v1Nh3')                    
                    all_likes_url=[]
                    for i in (likes):    
                        likes_text=i.find_element_by_tag_name('a').get_attribute('href')
                        all_likes_url.append(likes_text)
                    Total_likes=[]  
                    for i in all_likes_url:
                        driver.get(i)
                   
                        SCROLL_PAUSE_TIME = 0.5
                        last_height = driver.execute_script("return document.body.scrollHeight")
                        
                        while True:
                            # Scroll down to bottom
                            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")                        
                            # Wait to load page
                            time.sleep(SCROLL_PAUSE_TIME)                        
                            # Calculate new scroll height and compare with last scroll height
                            new_height = driver.execute_script("return document.body.scrollHeight")
                            if new_height == last_height:
                                break
                            last_height = new_height
                        time.sleep(5)
                        try:
                            likes=driver.find_element_by_css_selector('#react-root > section > main > div > div > article > div.eo2As > section.EDfFK.ygqzn > div > div > button')
                        
                            likes_count_text=likes.text  
                                                                          
                            likes_count=likes_count_text.replace('likes','')
                            Total_likes.append(int(likes_count))                        
                       
                        except:
                            print('e') 
                    max_likes.append(max(Total_likes))'''
                   
            except:
                    print('e')
                    
                    Total_posts.append('null') 
                    Total_followers.append('null') 
                    Total_following.append('null') 
                    #max_likes.append('null')


rows = zip(Total_posts,Total_followers,Total_following)
with open('rankinginsta.csv', "w",newline='') as f:
    writer = csv.writer(f)
    rowheader = ["Posts","Followers","Following"]
    writer.writerow(rowheader) 
    for row in rows:
        writer.writerow(row)


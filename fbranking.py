import openpyxl
from selenium import webdriver
import time
from statistics import mean
import xlwt
from tempfile import TemporaryFile
import csv


start_time=time.time()
path = "F:\\erin\\urls_fb.xlsx"
wb_obj = openpyxl.load_workbook(path) 
sheet_obj = wb_obj.active 
m_row = sheet_obj.max_row 

chrome_options = webdriver.ChromeOptions()

time.sleep(5)
prefs = {"profile.default_content_setting_values.notifications" : 2}
chrome_options.add_experimental_option("prefs",prefs)
driver = webdriver.Chrome(chrome_options=chrome_options)
driver.get("https://www.facebook.com")
#login
login=driver.find_element_by_css_selector('#email')
pass_word=driver.find_element_by_css_selector('#pass')
login_button=driver.find_element_by_css_selector('#u_0_2')
login.send_keys('adwitiya.kr@gmail.com')
pass_word.send_keys('Sum#1234')                                              
login_button.click()
 

num_likes_count=[]  
num_followers_count=[]
max_likes_count=[]   
min_likes_count=[] 
mean_likes_count=[]
rating_count_all=[]
rated_by_count=[]

 
for i in range(180,220):# m_row + 1): 
            
            cell_obj = sheet_obj.cell(row = i, column = 2)
            
            
            try:
                
                fb_url=cell_obj.value            
                
                if 'null' in fb_url:
                    num_likes_count.append(0)
                    num_followers_count.append(0)
                    max_likes_count.append(0)
                    min_likes_count.append(0)
                    mean_likes_count.append(0)
                    rating_count_all.append(0)
                    rated_by_count.append(0)
                    
                else:
                    
                    driver.get(fb_url)
                    
                    
                    
                    time.sleep(10)
                    try:
                        x=driver.find_element_by_class_name('close')
                        x.click()
                    except:
                        print('no close')
                    time.sleep(10)
                    try:
                        rating=driver.find_element_by_xpath('//*[@id="PagesProfileHomeSecondaryColumnPagelet"]/div/div[1]/div/a/div/div[2]/div/span[1]')
                        rating_text=rating.text
                        if 'out of 5' in rating_text:
                            rating_text=rating_text.replace('out of 5','')
                        rating_count_all.append(rating_text)
                        name=driver.find_elements_by_class_name('_2w0b')
                        popu=name[0].text
                        popu=popu.split(sep=' ')
                        total_people=popu[-2]
                        rated_by_count.append(total_people)
                        print('y')
                    except:
                        print('No rating')
                        rating_count_all.append(0)
                        rated_by_count.append(0)
                    '''  
                    try:
                        likes=driver.find_element_by_css_selector('#PagesProfileHomeSecondaryColumnPagelet > div > div:nth-child(2) > div > div._4-u2._6590._3xaf._4-u8 > div:nth-child(2) > div > div._4bl9 > div')
                        likes_text=likes.text               
                        likes_count=likes_text.replace('people like this','')
                        num_likes_count.append(likes_count)
                    except:
                        num_likes_count.append(0)
                    try:
                        follow=driver.find_element_by_css_selector('#PagesProfileHomeSecondaryColumnPagelet > div > div:nth-child(2) > div > div._4-u2._6590._3xaf._4-u8 > div:nth-child(3) > div > div._4bl9 > div')
                        follow_text=follow.text
                        follow_count=follow_text.replace('people follow this','')
                        num_followers_count.append(follow_count)
                    except:
                        num_followers_count.append(0)
                        
                    '''
                    try:     
                        page_like_follow=driver.find_elements_by_class_name('_4bl9') #likes and follow
                        page_like_follow_count=[]
                        for i in (page_like_follow):
                            page_like_follow_count.append(i.text)
                        for i in page_like_follow_count:
                            if 'people like' in i:
                                num_likes=i.replace('people like this','').strip()
                                num_likes_count.append(num_likes)
                            elif 'people follow' in i:    
                                num_follow=i.replace('people follow this','').strip()
                                num_followers_count.append(num_follow) 
                        print('z')
                    except:
                        num_likes_count.append(0)
                        num_followers_count.append(0)
                    
                   
                    
                    SCROLL_PAUSE_TIME = 0.5
                    last_height = driver.execute_script("return document.body.scrollHeight")                    
                    while True:
                        # Scroll down to bottom
                        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    
                        # Wait to load page
                        time.sleep(SCROLL_PAUSE_TIME)
                    
                        # Calculate new scroll height and compare with last scroll height
                        new_height = driver.execute_script("return document.body.scrollHeight")
                        if new_height == last_height:
                            break
                        last_height = new_height
                    time.sleep(10)  
                    
                    try:
                        name=driver.find_elements_by_class_name('_81hb')
                        name_text=[]
                        for i in (name):
                            name_text.append(i.text)
                        name_text=list(map(int,name_text))
                        min_like=min(name_text)
                        max_like=max(name_text)
                        mean_like=mean(name_text)
                        max_likes_count.append(max_like)
                        min_likes_count.append(min_like)
                        mean_likes_count.append(mean_like)
                        
                    except:
                        max_likes_count.append(0)
                        min_likes_count.append(0)
                        mean_likes_count.append(0)
                        
                       
                  
                       
                    
                
            except:
                print('error')
                
                
                num_likes_count.append('null') 
                num_followers_count.append('null') 
                
                max_likes_count.append('null') 
                min_likes_count.append('null')
                mean_likes_count.append('null')
                rating_count_all.append('null')
                rated_by_count.append('null')
               
                      
                


rows = zip(num_likes_count,num_followers_count,min_likes_count,max_likes_count,mean_likes_count,rating_count_all,rated_by_count)
with open('rankingfb.csv', "w",newline='') as f:
    writer = csv.writer(f)
    rowheader = ["Likes","Followers","Min_Likes","Max_Likes","Mean_likes","Rating","Rated_by"]
    writer.writerow(rowheader) 
    for row in rows:
        writer.writerow(row)
              
                
                
                
                
                
                
                
                
                